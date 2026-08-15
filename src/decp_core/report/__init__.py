"""报告导出：HTML 报告 + Excel 报表。

面向产品经理 / 需求收集人员：在数字员工 agent 中通过自然语言指令查看、下载结果。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from jinja2 import Environment, PackageLoader

from decp_core.models import AnalysisResult, Feedback, Requirement


class ReportService:
    """生成需求分析报告（HTML）与需求清单（Excel）。"""

    def __init__(self, out_dir: str | Path) -> None:
        self.out_dir = Path(out_dir)
        self.out_dir.mkdir(parents=True, exist_ok=True)

    # ---- HTML ----
    async def build_html_report(
        self,
        feedbacks: list[Feedback],
        requirements: list[Requirement],
        analysis: AnalysisResult,
        title: str = "产品需求收集、整理与分析报告",
    ) -> Path:
        env = Environment(loader=PackageLoader("decp_core", "report/templates"))
        tpl = env.get_template("report.html")
        html = tpl.render(
            title=title,
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            stats={
                "feedback_total": len(feedbacks),
                "requirement_total": len(requirements),
                "cluster_count": len(analysis.clusters),
                "duplicate_groups": len(analysis.duplicate_groups),
            },
            categories=analysis.categories,
            clusters=analysis.clusters,
            priorities=analysis.priorities,
            impact=analysis.impact,
            duplicates=analysis.duplicate_groups,
            sources=analysis.sources_verified,
            requirements=requirements,
            feedbacks=feedbacks,
        )
        path = self.out_dir / f"requirement-analysis-{datetime.now():%Y%m%d-%H%M%S}.html"
        path.write_text(html, encoding="utf-8")
        return path

    # ---- Excel ----
    async def build_excel_report(
        self,
        requirements: list[Requirement],
        feedbacks: list[Feedback],
        analysis: AnalysisResult,
    ) -> Path:
        wb = openpyxl.Workbook()
        # Sheet1 需求清单
        ws = wb.active
        ws.title = "需求清单"
        headers = [
            "需求ID", "标题", "模块", "优先级", "状态", "版本",
            "影响客户数", "相似反馈数", "置信度", "关联反馈数", "创建时间", "更新时间",
        ]
        self._write_headers(ws, headers)
        for r in requirements:
            ws.append([
                r.id, r.title, r.module or "", r.priority, r.status, r.version,
                r.impact_customers, r.similar_feedback_count, r.confidence,
                len(r.feedback_ids), r.created_at.strftime("%Y-%m-%d %H:%M"),
                r.updated_at.strftime("%Y-%m-%d %H:%M"),
            ])
        self._auto_width(ws, len(headers))

        # Sheet2 反馈明细
        ws2 = wb.create_sheet("反馈明细")
        h2 = ["反馈ID", "客户", "模块", "类型", "影响", "渠道", "提交人", "内容", "优先级建议"]
        self._write_headers(ws2, h2)
        for f in feedbacks:
            ws2.append([
                f.id, f.customer or "", f.module or "",
                f.structured.get("feedback_type", ""),
                f.structured.get("impact_severity", ""),
                f.channel, f.submitted_by, f.content,
                analysis.priorities.get(f.id, ""),
            ])
        self._auto_width(ws2, len(h2))

        # Sheet3 聚类
        ws3 = wb.create_sheet("聚类分析")
        h3 = ["聚类ID", "主题", "反馈数", "覆盖反馈"]
        self._write_headers(ws3, h3)
        for c in analysis.clusters:
            ws3.append([c["id"], c["title"], c["count"], "\n".join(c["feedback_ids"])])
        self._auto_width(ws3, len(h3))

        path = self.out_dir / f"requirement-list-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
        wb.save(str(path))
        return path

    # ---- helpers ----
    @staticmethod
    def _write_headers(ws: Any, headers: list[str]) -> None:
        fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill

    @staticmethod
    def _auto_width(ws: Any, ncols: int) -> None:
        for i in range(1, ncols + 1):
            letter = get_column_letter(i)
            max_len = 10
            for row in ws.iter_rows(min_col=i, max_col=i):
                for cell in row:
                    v = cell.value
                    if v is not None:
                        max_len = max(max_len, min(len(str(v)), 60))
            ws.column_dimensions[letter].width = min(max_len + 2, 80)
